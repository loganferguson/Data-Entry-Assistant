from tkinter import *
from tkinter import ttk
import tkinter.font as Font
import os
from help import commands
from dotenv import find_dotenv, load_dotenv
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait

root = Tk()
root.title("Space Data Entry Assistant")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_pos = (screen_width - 1000) // 2
root.geometry(f'1000x100+{x_pos}+{screen_height - 200}')
root.configure(bg="gray35")
style = ttk.Style()

dotenv_path = find_dotenv()
load_dotenv(dotenv_path)
USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")
url = "https://caits.iu.edu/Space/Spaces"
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver.set_window_position(0, 0)
driver.set_window_size(screen_width - 5, screen_height - 200)
driver.get(url)

option_commands = ["openxl", "lr", "dup", "ed", "name", "delp", "note", "setnote", "org", "rm", "bc", "type", "station", "setbc", "edp", "p", "rd", "add", "end", "bld", "space", "apc", "bca", "rma", "det", "eda", "start"]
person_options = []
current_person_index = 0
in_dept_mode = True
xl_fullpath = ""
xl_file = ""
xl_file_name = ""
xl_sheet = ""
active_row = 0
xl_org_code = ""
xl_building_code = ""
xl_room_number = ""
xl_space_type = ""
xl_update_column = ""
default_note = "Update occupant per DMFC inventory 7/16/2025 - LF"
default_note_2 = "This is the default 2 note"


def LoadSheet(filepath):
    global filename
    global xl_fullpath   
    filename = os.listdir(filepath)[0]
    file_name_label.configure(text=f"File: {filename}")
    file_name_label.pack(side=TOP, anchor='nw')
    xl_fullpath = os.path.join(filepath, filename)
    global xl_file
    xl_file = openpyxl.load_workbook(xl_fullpath)
    global xl_sheet
    xl_sheet = xl_file["Space Inventory"]
    active_row_label.pack_forget()

def LoadRow(row):
    global active_row
    active_row = row
    global xl_org_code
    xl_org_code = xl_sheet.cell(row=active_row, column=3).value if not in_dept_mode else xl_sheet.cell(row=active_row, column=24).value
    if not in_dept_mode:
        global xl_building_code
        xl_building_code = xl_sheet.cell(row=active_row, column=5).value
    global xl_room_number
    xl_room_number = xl_sheet.cell(row=active_row, column=7).value if not in_dept_mode else xl_sheet.cell(row=active_row, column=4).value
    global xl_space_type
    xl_space_type = xl_sheet.cell(row=active_row, column=12).value if not in_dept_mode else xl_sheet.cell(row=active_row, column=11).value
    global xl_update_column
    xl_update_column = xl_sheet.cell(row=active_row, column=15).value if not in_dept_mode else xl_sheet.cell(row=active_row, column=24).value
    
    if(xl_building_code is not None and xl_room_number is not None):
        for row in xl_table.get_children():
            xl_table.delete(row)
        xl_table.insert(parent='', index=0, values=(xl_org_code, xl_building_code, xl_room_number, xl_space_type, xl_update_column))
    active_row_label.configure(text=f"Row: {active_row}")
    active_row_label.pack(side=TOP, anchor='nw')
    file_name_label.pack(side=RIGHT, anchor='nw')

def LoadCurrentRow():
    global active_row
    LoadRow(active_row)

def MarkRowComplete():
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
    global active_row
    global xl_sheet
    current_row_fill_string = str(active_row) + ':' + str(active_row)
    for cell in xl_sheet[current_row_fill_string]:
        cell.fill = yellow_fill
    # global xl_update_column
    # xl_update_column = xl_update_column + " LF" 
    # xl_sheet.cell(row=active_row, column=44).value = xl_update_column

def SaveXLFile():
    global xl_fullpath
    xl_file.save(xl_fullpath)

def CreateNew():
    create_space_button = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/section/div/p/a')
    create_space_button.click()

def Select_New_Space_Building(building_code):
    building_dropdown = driver.find_element(By.ID, 'Building')
    select_object = Select(building_dropdown)
    for option in select_object.options:
        if building_code in option.text:
            select_object.select_by_visible_text(option.text)
            break

def Select_Space(room):
    composite = 'Health Information & Translational Sciences / ' + room
    space_dropdown = driver.find_element(By.ID, 'Space')
    select_object = Select(space_dropdown)
    for option in select_object.options[1000:]:
        if composite in option.text:
            select_object.select_by_visible_text(option.text)
            break
    
def Set_Assignment_Percentage(percentage):
    percentage_input = driver.find_element(By.ID, "Percentage")
    percentage_input.clear()
    percentage_input.send_keys(percentage)
def SearchSpace():
    if xl_building_code == "" or xl_room_number == "":
        search_space_btn = driver.find_element(By.XPATH, '//*[@id="Body-anchor"]/form/div[1]/div[3]/div/div[2]/input')
        search_space_btn.click()
        driver.implicitly_wait(2)
        return
    building_code_input = driver.find_element(By.XPATH, '//*[@id="SelectedBldgCode"]')
    room_number_input = driver.find_element(By.XPATH, '//*[@id="SelectedRoomNumber"]')
    building_code_input.clear()
    room_number_input.clear()
    building_code_input.send_keys(xl_building_code)
    room_number_input.send_keys(xl_room_number)
    search_space_btn = driver.find_element(By.XPATH, '//*[@id="Body-anchor"]/form/div[1]/div[3]/div/div[2]/input')
    search_space_btn.click()
    driver.implicitly_wait(2)
    
def ViewAssignment():
    assignment_btn = driver.find_element(By.XPATH, '//*[@id="Body-anchor"]/form/div[2]/table/tbody/tr/td[7]/a[1]')
    assignment_btn.click()
    driver.implicitly_wait(4)

def DuplicateAssignment(assignment):
    assignment_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = assignment_table.find_element(By.TAG_NAME, 'tbody')
    assignment_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = assignment_rows[(assignment - 1)]
    duplicate_btn = selected_row.find_element(By.CLASS_NAME, 'duplicate-link')
    duplicate_btn.click()

def Edit(assignment):
    assignment_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = assignment_table.find_element(By.TAG_NAME, 'tbody')
    assignment_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = assignment_rows[(assignment - 1)]
    row_options = selected_row.find_elements(By.TAG_NAME, 'a')
    edit_btn = row_options[0]
    edit_btn.click()

def EditActive(assignment):
    assignment_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = assignment_table.find_element(By.TAG_NAME, 'tbody')
    assignment_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = assignment_rows[(assignment - 1)]
    row_options = selected_row.find_elements(By.TAG_NAME, 'a')
    edit_btn = row_options[1]
    edit_btn.click()

def Details(assignment):
    assignment_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = assignment_table.find_element(By.TAG_NAME, 'tbody')
    assignment_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = assignment_rows[(assignment - 1)]
    row_options = selected_row.find_elements(By.TAG_NAME, 'a')
    details_btn = row_options[len(row_options) - 1]
    details_btn.click()
    
def AcceptAlert():
    Alert(driver).accept()

def AddPerson():
    add_person_btn = driver.find_element(By.CSS_SELECTOR, "input[value='Add Person']")
    add_person_btn.click()

def CreatePerson():
    create_btn = driver.find_element(By.CSS_SELECTOR, "input[value='Create']")
    create_btn.click()

def SearchSpaceManual():
    search_btn = driver.find_element(By.CSS_SELECTOR, "input[value='Search']")
    search_btn.click()

def EditPerson(person):
    person_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = person_table.find_element(By.TAG_NAME, 'tbody')
    person_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = person_rows[(person - 1)]
    edit_btn = selected_row.find_element(By.CSS_SELECTOR, "input[value='Edit']")
    edit_btn.click()

def EditPersonPercentage(percentage):
    percentage_input = driver.find_element(By.ID, 'AssignmentPercentage')
    percentage_input.clear()
    percentage_input.send_keys(percentage)

def EnterName(name: str):
    name_input = driver.find_element(By.ID, 'PersonName')
    name_input.clear()
    name_input.send_keys(name)

def DeletePerson(person):
    person_table = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/section/div/form/div/div[7]/div[2]/table')
    table_body = person_table.find_element(By.TAG_NAME, 'tbody')
    person_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = person_rows[person - 1]
    person_options = selected_row.find_elements(By.TAG_NAME, 'td')[-1]
    delete_btn = person_options.find_elements(By.TAG_NAME, 'input')[-1]
    delete_btn.click()

def ConfirmDeletePerson():
    delete_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/section/div/div/form/div/input')
    delete_btn.click()

def AddOrg(org_code: str):
    org_dropdown = driver.find_element(By.ID, 'Organization')
    select_object = Select(org_dropdown)
    for option in select_object.options:
        if org_code in option.text:
            select_object.select_by_visible_text(option.text)
            break

def AddSpaceType(space_type: str):
    spacetype_dropdown = driver.find_element(By.ID, 'SpaceType')
    select_object = Select(spacetype_dropdown)
    for option in select_object.options:
        if space_type in option.text:
            select_object.select_by_visible_text(option.text)
            break

def AddRoomDescription(description):
    description_text_area = driver.find_element(By.ID, 'RoomDescription')
    description_text_area.clear()
    description_text_area.send_keys(description)

def AddStation(num: int):
    station_input = driver.find_element(By.ID, 'Stations')
    station_input.clear()
    station_input.send_keys(num)

def EnterBuilding(building_code: str):
    building_code_input = driver.find_element(By.ID, 'SelectedBldgCode')
    building_code_input.clear()
    building_code_input.send_keys(building_code)

def EnterAssignmentBuildingCode(building_code: str):
    building_code_input = driver.find_element(By.ID, 'BldgCode')
    building_code_input.clear()
    building_code_input.send_keys(building_code)

def SetDefaultBuilding(building_code: str):
    global xl_building_code
    xl_building_code = building_code
    LoadCurrentRow()

def EnterRoom(room: str):
    room_input = driver.find_element(By.ID, 'SelectedRoomNumber')
    room_input.clear()
    room_input.send_keys(room)

def EnterAssignmentRoom(room: str):
    room_input = driver.find_element(By.ID, 'RoomNumber')
    room_input.clear()
    room_input.send_keys(room)

def AddNote(note):
    note_text_area = driver.find_element(By.ID, 'Notes')
    note_text_area.clear()
    global default_note
    global default_note_2
    if(note == "def"):
        note_text_area.send_keys(default_note)
    elif(note == "def2"):
        note_text_area.send_keys(default_note_2)
    else:
        note_text_area.send_keys(note)

def SetDefaultNote(new_note):
    global default_note
    default_note = new_note

def SetDefaultNote_2(new_note):
    global default_note_2
    default_note_2 = new_note

def CycleCurrentCheckbox():
    current_assignments_checkbox = driver.find_element(By.ID, 'CurrentAssignments')
    current_assignments_checkbox.click()

def CycleUpdatePreviousEnd():
    update_previous_end_checkbox = driver.find_element(By.ID, 'PreviousEndDateFlag')
    update_previous_end_checkbox.click()

def AddStartDate(startdate):
    date_input = driver.find_element(By.ID, 'StartDate')
    date_input.clear()
    date_input.send_keys(startdate)

def AddEndDate(enddate):
    date_input = driver.find_element(By.ID, 'EndDate')
    date_input.clear()
    date_input.send_keys(enddate)

def Update():
    update_btn = driver.find_element(By.CSS_SELECTOR,  "input[value='Update']")
    update_btn.click()

def Clear():
    clear_btn = driver.find_element(By.CSS_SELECTOR,  "input[value='Clear']")
    clear_btn.click()

def AddAssignment(responsible):
    assignment_table = driver.find_element(By.TAG_NAME, 'table')
    table_body = assignment_table.find_element(By.TAG_NAME, 'tbody')
    assignment_rows = table_body.find_elements(By.TAG_NAME, 'tr')
    selected_row = assignment_rows[(responsible - 1)]
    row_options = selected_row.find_elements(By.TAG_NAME, 'a')
    add_btn = row_options[0]
    add_btn.click()


def Login():
    username_input = driver.find_element(By.XPATH, '//*[@id="username"]')
    password_input = driver.find_element(By.XPATH, '//*[@id="password"]')
    username_input.send_keys(USERNAME)
    password_input.send_keys(PASSWORD)
    login_button = driver.find_element(By.XPATH, '//*[@id="login-button"]')
    login_button.click()

def YesMyDevice():
    yes_btn = driver.find_element(By.XPATH, '/html/body/div/div/div[1]/div/div[2]/div[3]/button')
    yes_btn.click()

def NavigateMenu(menu_url):
    driver.get(menu_url)

def NextRow():
    global active_row
    active_row += 1
    LoadRow(active_row)
    active_row_label.configure(text=f"Row: {active_row}")
    active_row_label.pack(side=TOP, anchor='nw')

def PreviousRow():
    global active_row
    active_row -= 1
    LoadRow(active_row)
    active_row_label.configure(text=f"Row: {active_row}")
    active_row_label.pack(side=TOP, anchor='nw')

def ViewCommands():
    help_window = Toplevel(root)
    help_window.title("View Commands")
    help_window.geometry("500x800")
    help_window.anchor("w")
    commands_label = Label(help_window, text=commands, bg="gray35", fg="OliveDrab1")
    commands_label.pack(side=LEFT, anchor="w", )
def ManageCommand(_):
    last_command = cmd_box.get()
    command = last_command.split()[0]
    command_option = ""
    if (command in option_commands):
        command_option = last_command.split(' ', 1)[1]
    cmd_box.delete(0, END)
    match command:
        case 'li':
            Login()
        case 'ymd':
            YesMyDevice()
        case 'openxl':
            LoadSheet(command_option)
        case 'lr':
            LoadRow(int(command_option))
        case 'nr':
            NextRow()
        case 'pr':
            PreviousRow()
        case 'setbc':
            SetDefaultBuilding(command_option)
        case 'sp':
            NavigateMenu('https://caits.iu.edu/Space/Spaces/Index/28')
        case 'ass':
            NavigateMenu('https://caits.iu.edu/Space/Assignments/Index/23')
        case 'back':
            driver.back()
        case 'bc':
            EnterBuilding(command_option)
        case 'bca':
            EnterAssignmentBuildingCode(command_option)
        case 'rma':
            EnterAssignmentRoom(command_option)
        case 'rm':
            EnterRoom(command_option)
        case 'ss':
            SearchSpace()
        case 'new':
            CreateNew()
        case 'bld':
            Select_New_Space_Building(command_option)
        case 'space':
            Select_Space(command_option)
        case 'apc':
            Set_Assignment_Percentage(command_option)
        case 'sr':
            SearchSpaceManual()
        case 'va':
            ViewAssignment()
        case 'ca':
            CycleCurrentCheckbox()
        case 'upe':
            CycleUpdatePreviousEnd()
        case 'dup':
            DuplicateAssignment(int(command_option))
        case 'ed':
            Edit(int(command_option))
        case 'eda':
            EditActive(int(command_option))
        case 'add':
            AddAssignment(int(command_option))
        case 'det':
            Details(int(command_option))
        case 'ok':
            AcceptAlert()
        case 'ap':
            AddPerson()
        case 'edp':
            EditPerson(int(command_option))
        case 'p':
            EditPersonPercentage(command_option)
        case 'cr':
            CreatePerson()
        case 'org':
            AddOrg(command_option)
        case 'type':
            AddSpaceType(command_option)
        case 'rd':
            AddRoomDescription(command_option)
        case 'station':
            AddStation(command_option)
        case 'name':
            EnterName(command_option)
        case 'delp':
            DeletePerson(int(command_option))
        case 'cdelp':
            ConfirmDeletePerson()
        case 'note':
            AddNote(command_option)
        case 'setnote':
            SetDefaultNote(command_option)
        case 'setnote2':
            SetDefaultNote_2(command_option)
        case 'start':
            AddStartDate(command_option)
        case 'end':
            AddEndDate(command_option)
        case 'u':
            Update()
        case 'clr':
            Clear()
        case 'lcr':
            LoadCurrentRow()
        case 'done':
            MarkRowComplete()
        case 'savexl':
            SaveXLFile()
        case 'help':
            ViewCommands()


#----------------------------COMMAND LINE----------------------------#
custom_font = Font.Font(family="Consolas", size=12)
cmd_box = Entry(root, bg="gray1", fg="OliveDrab1", width=60, font=custom_font)
cmd_box.config(insertbackground="OliveDrab1", insertwidth=10)

cmd_box.bind("<Return>", ManageCommand)
cmd_box.pack(side=TOP, anchor='nw', padx=2, pady=2)
#-----------------------------XL ROW DATA SECTION------------------------------#
row_data_frame = Frame(root, bg="gray35")
row_data_frame.pack(side=LEFT, anchor="nw", padx=2)

#-----------------XL TREE VIEW--------------#
style.theme_use("alt")
style.map('Treeview.Heading',
          background=[('active', 'gray35')])
style.map('Treeview',
          background=[('selected', 'gray30')],
          foreground=[('selected', 'OliveDrab2')])
style.configure("My.Treeview",
                background="gray35",
                foreground="OliveDrab2",
                fieldbackground="gray35")
style.configure("My.Treeview.Heading",
                background="gray35",
                foreground="CadetBlue1",
                font=("Calibri Bold", 10, "bold")
                )
xl_table = ttk.Treeview(row_data_frame, columns=('1', '2', '3', '4', '5'), show='headings', height=1, style="My.Treeview")
xl_table.heading('1', text="Assigned Org")
xl_table.heading('2', text="Building Code")
xl_table.heading('3', text="Room")
xl_table.heading('4', text="Space Type")
xl_table.heading('5', text="Update")
xl_table.pack(side=TOP, anchor='nw')
active_row_label = Label(row_data_frame, text="Row:", padx=2, pady=2,bg="gray35", fg="light pink", font=custom_font)
file_name_label = Label(row_data_frame, text="File:", padx=2, pady=2,bg="gray35", fg="goldenrod", font=custom_font)
cmd_box.focus_set()
root.mainloop()
